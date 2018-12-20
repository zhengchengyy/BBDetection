from openpyxl import load_workbook
import time, datetime

workbook = load_workbook('heartRate.xlsx')
sheets = workbook.get_sheet_names()  # 从名称获取sheet
booksheet = workbook.get_sheet_by_name(sheets[0])  #取第一张表

def timeToSecond(t):
	return t.hour*3600+t.minute*60+t.second

# 获取第一列的时间和第二列的心率
time = []
rate = []
for row in range(2,booksheet.max_row+1):
	t = booksheet.cell(row, 1).value
	sec = timeToSecond(t)
	time.append(sec)

	r = booksheet.cell(row, 2).value
	rate.append(r)

print(rate)
# print(type(time[1].hour))
# print(time[1].hour)
# print(time[1].minute)
# print(time[1].second)

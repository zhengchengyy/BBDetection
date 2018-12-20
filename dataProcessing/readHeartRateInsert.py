from openpyxl import load_workbook
from openpyxl import Workbook
import time, datetime
from openpyxl.styles import Alignment

inputfilename = 'heartrate.xlsx'
outputfilename = 'easy.xlsx'

workbook = load_workbook(inputfilename)
sheets = workbook.sheetnames  # 从名称获取sheet
booksheet = workbook[sheets[0]] #取第一张表

# transform hh:mm:ss to s
def timeToSecond(t):
	return t.hour*3600+t.minute*60+t.second

# 获取第一列的时间和第二列的心率
time = []
rate = []
for row in range(2,booksheet.max_row+1):
	t = booksheet.cell(row, 1).value
	sec = timeToSecond(t)
	time.append(sec)

	r = booksheet.cell(row, 2).value
	rate.append(r)

# 定义一个全天时间序列
all_time = range(1, 3600*24+1)

def getNeighbour(all_time, time, rate, i, index):
    if index == 0:
        return rate[0]
    elif (all_time[i]-time[index - 1]) <= (time[index]-all_time[i]):
        return rate[i - 1]
    else:
        return rate[i]

def insertValue(all_time, time):
    index = 0
    for i in range(len(all_time)):
        if index == len(time) and i == len(all_time):
            break
        elif index == len(time):
	        rate.insert(i,rate[i-1])
        elif time[index] == all_time[i]:
            index = index+1
        else:
	        rate.insert(i, getNeighbour(all_time, time, rate, i, index))

# transform s to hh:mm:ss
def secondToTime(sec):
	m, s = divmod(sec, 60)
	h, m = divmod(m, 60)
	str = "%02d:%02d:%02d" % (h, m, s)
	return str

insertValue(all_time, time)

workbook2 = load_workbook(outputfilename)
sheets2 = workbook2.sheetnames  # 从名称获取sheet
booksheet2 = workbook2[sheets2[0]]  #取第一张表
# set column width
col_list = ['A', 'B', 'C', 'D', 'E', 'F']
for col in col_list:
	booksheet2.column_dimensions[col].width = 20

# input parameter timeStamp must be int type
# The preceding variables time lead to ambiguity(歧义), so import time
import time
def timeStampToTime(timeStamp):
	# because timeStamp include decimal,so rounding
	if timeStamp-int(timeStamp)>=0.5:
		timeStamp = timeStamp + 1
	timeArray = time.localtime(timeStamp)
	otherStyleTime = time.strftime("%H:%M:%S", timeArray)
	return otherStyleTime

# read timeStamp from xlsx
volt_all_time = []
for row in range(2,booksheet2.max_row+1):
	timestamp = booksheet2.cell(row, 1).value
	stdTime = timeStampToTime(float(timestamp))
	volt_all_time.append(stdTime)

# add standard time to new table
booksheet2.cell(1, 5).value = 'timetostd'
for i in range(len(volt_all_time)):
	booksheet2.cell(i + 2, 5).value = str(volt_all_time[i])

# add rate to new table
booksheet2.cell(1, 6).value = 'heartrate'
i = 0
j = 0
while i!= len(volt_all_time):
	if volt_all_time[i] == secondToTime(all_time[j]):
		booksheet2.cell(i + 2, 6).value = str(rate[j])
		i = i + 1
	else:
		j = j + 1

workbook2.save(outputfilename)
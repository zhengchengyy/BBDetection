# 从mongodb读取原始数据，用openpyxl写入表格
# （xlrd和xlwt处理的是xls文件，单个sheet最大行数是65535,openpyxl函数最大行数达到1048576）
# 对表格中三块板子的电压数据进行数据对齐（如果其中一块板子没有相应时间点的电压值，就对其就近插值处理）

from pymongo import MongoClient
from exceptions import *
from openpyxl import Workbook

config = {'action':'turn_over',
          'db':'beaglebone',
          'tag_collection':'tags_411',
          'volt_collection':'volts_411'}

# 合并两个设备的有序时间列表
def merge(time1,time2):
    res_time = []
    p = 0
    q = 0
    while p!=len(time1) and q!=len(time2):
        if time1[p]==time2[q]:
            res_time.append(time1[p])
            p = p+1
            q = q+1
        elif time1[p]<time2[q]:
            res_time.append(time1[p])
            p = p+1
        else:
            res_time.append(time2[q])
            q = q+1

    while p!=len(time1):
        res_time.append(time1[p])
        p = p + 1

    while q!=len(time2):
        res_time.append(time2[q])
        q = q + 1
    return res_time

# 获取邻近时间值，如果相邻位置距离相同，则选择左边
# all_time：所有时间列表；time:某设备时间列表；volt:某设备电压列表；i:all_time的索引；index:time的索引
def getNeighbour(all_time, time, volt, i, index):
    if index == 0:
        return volt[0]
    elif (all_time[i]-time[index - 1]) <= (time[index]-all_time[i]):
        return volt[i - 1]
    else:
        return volt[i]

# 对所在时间内不存在的电压进行插值处理，参数说明和上面一样
def insertValue(all_time, times, device_no ,volts):
    index = 0
    for i in range(len(all_time)):
        if index == len(times[device_no]) and i == len(all_time):
            break
        elif index == len(times[device_no]):
            volts[device_no].insert(i,volts[device_no][i-1])
        elif times[device_no][index] == all_time[i]:
            index = index+1
        else:
            volts[device_no].insert(i, getNeighbour(all_time, times[device_no], volts[device_no], i, index))


def save_to_excel(action, db, volt_collection, tag_collection,port=27017, host='localhost', ndevices=3):
    client = MongoClient(port=port, host=host)
    database = client[db]
    tag_collection = database[tag_collection]
    volt_collection = database[volt_collection]

    if volt_collection.count_documents({}) + tag_collection.count_documents({}) < 2:
        raise CollectionError('Collection not found, please check names of the collection and database')

    # save to excel
    workbook = Workbook()
    booksheet = workbook.active  # 获取当前活跃的sheet,默认是第一个sheet
    booksheet.title = action
    booksheet['A1'] = 'time'
    booksheet['B1'] = 'volt_1'
    booksheet['C1'] = 'volt_2'
    booksheet['D1'] = 'volt_3'

    for tag in tag_collection.find({'tag': action}):
        inittime, termtime = tag['inittime'], tag['termtime']

        # get the arrays according to which we will plot later
        times, volts = {}, {}
        for i in range(1, ndevices + 1):
            times[i] = []
            volts[i] = []

        for volt in volt_collection.find({'time': {'$gt': inittime,'$lt': termtime}}):
            device_no = int(volt['device_no'])
            v = volt['voltage']
            time = volt['time']
            times[device_no].append(time)
            volts[device_no].append(v)
            # print(times[device_no])

        # data synchronization
        all_time = {}
        all_time = merge(merge(times[1], times[2]), times[3])
        for device_no in range(1, ndevices + 1):
            insertValue(all_time, times, device_no, volts)

        # save data
        for i in range(len(all_time)):
            booksheet.cell(i + 2, 1).value = str(all_time[i])

        for col in range(1, ndevices + 1):
            for row in range(len(volts[col])):
                booksheet.cell(row + 2, col + 1).value = str(volts[col][row])

        workbook.save(action + "_syn" + '.xlsx')


save_to_excel(action=config['action'],
              db=config['db'],
              tag_collection=config['tag_collection'],
              volt_collection=config['volt_collection'])
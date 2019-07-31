# 从mongodb读取数据，用openpyxl写入表格
# （xlrd和xlwt处理的是xls文件，单个sheet最大行数是65535,openpyxl函数最大行数达到1048576）

from pymongo import MongoClient
from exceptions import *
from openpyxl import Workbook

config = {'action':'turn_over',
          'db':'beaglebone',
          'tag_collection':'tags_411',
          'volt_collection':'volts_411'}

def save_to_excel(action, db, volt_collection, tag_collection,port=27017, host='localhost', ndevices=3):
    client = MongoClient(port=port, host=host)
    database = client[db]
    tag_collection = database[tag_collection]
    volt_collection = database[volt_collection]

    if volt_collection.count_documents({}) + tag_collection.count_documents({}) < 2:
        raise CollectionError('Collection not found, please check names of the collection and database')

    # save to excel
    workbook = Workbook()
    booksheet = workbook.active  # 获取当前活跃的sheet,默认是第一个sheet
    booksheet.title = action
    booksheet['A1'] = 'time_1'
    booksheet['B1'] = 'volt_1'
    booksheet['C1'] = 'time_2'
    booksheet['D1'] = 'volt_2'
    booksheet['E1'] = 'time_3'
    booksheet['F1'] = 'volt_3'
    col_list = ['A', 'B', 'C', 'D', 'E', 'F']
    for col in col_list:
        booksheet.column_dimensions[col].width = 25

    for tag in tag_collection.find({'tag': action}):
        inittime, termtime = tag['inittime'], tag['termtime']

        # get the arrays according to which we will plot later
        times, volts = {}, {}
        for i in range(1, ndevices + 1):
            times[i] = []
            volts[i] = []

        stable = 0.76
        for volt in volt_collection.find({'time': {'$gt': inittime,'$lt': termtime}}):
            device_no = int(volt['device_no'])
            time = volt['time']
            v = volt['voltage']
            times[device_no].append(time)
            volts[device_no].append(v)

        for col in range(1, ndevices*2 + 1):
            # 偶数列存入电压值
            if(col%2==0):
                c = (int)(col / 2)
                for row in range(len(volts[c])):
                    booksheet.cell(row + 2, col).value = volts[c][row]
            # 奇数列存入时间值
            else:
                c = (int)((col + 1) / 2)
                for row in range(len(times[c])):
                    booksheet.cell(row + 2, col).value = times[c][row]
        workbook.save(action + "_raw" + '.xlsx')


save_to_excel(action=config['action'],
              db=config['db'],
              tag_collection=config['tag_collection'],
              volt_collection=config['volt_collection'])